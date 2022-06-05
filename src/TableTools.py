import logging
import openpyxl
import datetime
import os
from collections import defaultdict
from calendar import Calendar
from typing import Dict, List, Union, Tuple
from pycel import ExcelCompiler
from openpyxl.utils import get_column_letter
from enum import Enum


SETTER_ROW_RANGE = (6, 60, 6)
SETTER_COLUMN = 1
GRADES_COLUMN = 2
SALARY_COLUMN = 17
RESULTS_COLUMN = 15
DATE_ROW = 4
CONTEST_NAME_ROW = 3
CONTEST_PRICE_ROW = 5

STARTS_FIRST_COL = 3
STARTS_SECOND_COL = 4
LAST_DATE_COL = 14

FIRST_CONTEST_COL = 13
SECOND_CONTEST_COL = 14

MONTHS = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul',
          8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
          }


class SetterStatus(Enum):
    ALREADY_EXISTS = 1
    ADDED = 2
    REMOVED = 3
    NOT_FOUND = 4
    FAILED_TO_GET = 5


class ContestStatus(Enum):
    ADDED = 1
    REMOVED = 2
    NO_PLACE = 3
    NO_CONTEST = 4
    FAILED_TO_GET = 5
    IS_CONTEST_SETTER = 6
    SETTER_ABSENT = 7


class ResultStatus(Enum):
    ADDED = 1
    FAIL_TO_GET_DATE = 2
    NOT_SETTER = 3


# logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)


class RoutesetterTable(object):
    def __init__(self, path=os.path.join('data', 'Routesetting.xlsx'), logger=None):
        self.tablepath = path
        self.logger = logger or logging.getLogger(__name__)
        self.__set_date()
        self.__open()
        self.logger.info(f'Bot Tabletools module launched: {self.current_date}')

    def __set_date(self) -> None:
        self.current_day = datetime.date.today().day
        self.current_month = datetime.date.today().month
        self.current_year = datetime.date.today().year
        self.current_date = datetime.date(self.current_year, self.current_month, self.current_day).strftime('%d.%m.%Y')
        self.next_month = self.current_month + 1
        self.next_year = self.current_year + 1
        self.year_to_use = self.current_year
        if self.current_month == 12:
            self.next_month = 1
            self.year_to_use = self.current_year + 1
        return

    def __open(self) -> None:
        try:
            self.workbook_r = openpyxl.load_workbook(self.tablepath, data_only=True)
            self.workbook_w = openpyxl.load_workbook(self.tablepath)
            self.workbook_f = ExcelCompiler(filename=self.tablepath)
            all_sheet = self.workbook_r.sheetnames
            self.current_month_name = ''.join([i for i in all_sheet if i.startswith(f'{str(self.current_year)}. {str(self.current_month)}')])
            if not self.current_month_name:
                self.new_sheet(self.current_month, self.current_year)
            self.current_month_sh_r = self.workbook_r[self.current_month_name]
            self.current_month_sh_w = self.workbook_w[self.current_month_name]
            if self.current_month_sh_r.cell(row=5, column=14).value:
                self.contests = 2
                self.contest_col = SECOND_CONTEST_COL
                return
            if self.current_month_sh_r.cell(row=5, column=13).value:
                self.contests = 1
                self.contest_col = FIRST_CONTEST_COL
                return
            self.contests = 0
        except Exception as e:
            self.logger.error(f'Error while excel opening(__open)\n{e}')
        return

    def get_table_as_bytes(self):
        return open(self.tablepath, 'rb')

    def get_cell_val(self, sheet_name: str, row_num: int, col_num: int) -> Union[str, int]:
        col = get_column_letter(col_num)
        return self.workbook_f.evaluate(f'{sheet_name}!{col}{str(row_num)}')

    def on_scheduler_new_day(self) -> None:
        self.__set_date()
        self.logger.info(f'New day had begun. Today is {self.current_date}, next month is {self.next_month}')

    def get_setting_dates(self, month=None, year=None) -> List[Tuple[int, int]]:
        month = month or self.next_month
        year = year or self.year_to_use
        tue_num, thu_num = 1, 3
        cal = Calendar().itermonthdays2(year, month)
        tue_thu = []
        for day in cal:
            if day[0] != 0 and day[1] in (tue_num, thu_num):
                tue_thu.append(day)
        return tue_thu

    def get_month_result(self) -> Dict[str, List[str]]:
        results = defaultdict(list)
        try:
            for row in range(*SETTER_ROW_RANGE):
                setter = self.current_month_sh_r.cell(row=row, column=SETTER_COLUMN).value
                if not setter:
                    continue
                for i in range(row, row + 6):
                    category = self.current_month_sh_r.cell(row=i, column=GRADES_COLUMN).value
                    amount = self.get_cell_val(self.current_month_name, i, RESULTS_COLUMN)
                    results[setter].append(f'{category}: {amount}')
            return results
        except Exception as e:
            self.logger.error(f'Error had happened while getting month results.\n{e}')
            return {}

    def get_user_result(self, username) -> List[str]:
        results = self.get_month_result()
        for key in results:
            if key == username:
                return results[username]
        return []

    def new_sheet(self, month=None, year=None) -> None:
        month = month or self.next_month
        year = year or self.year_to_use
        try:
            if f'{str(year)}. {str(month)}. {MONTHS[month]}' in self.workbook_r.sheetnames:
                self.logger.info('Tried to add new month but it already exists')
                return
            sheet = self.workbook_w['PATTERN']
            new_sh = self.workbook_w.copy_worksheet(sheet)
            new_sh.title = f'{str(year)}. {str(month)}. {MONTHS[month]}'
            dates = self.get_setting_dates(month=month, year=year)
            first = STARTS_FIRST_COL
            if dates[0][1] == 3:
                first = STARTS_SECOND_COL
            i = 0
            j = first
            while i != len(dates):
                date_str = datetime.date(year, month, dates[i][0]).strftime("%d.%m.%Y")
                new_sh.cell(row=DATE_ROW, column=j).value = date_str
                i += 1
                j += 1
            self.workbook_w.save(self.tablepath)
            self.logger.info('new_month succesfully added')
            self.__open()
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')

    def get_salary_info(self) -> Dict[str, int]:
        try:
            salaries = {}
            for row in range(*SETTER_ROW_RANGE):
                setter = self.current_month_sh_r.cell(row=row, column=SETTER_COLUMN).value
                if not setter:
                    continue
                salary = self.get_cell_val(self.current_month_name, row, SALARY_COLUMN)
                salaries[setter] = salary
            return salaries
        except Exception as e:
            self.logger.error(f'Error had happened getting salary info.\n{e}')
            return {}

    def richest_user(self) -> List[str]:
        salaries = self.get_salary_info()
        sorted_salaries = sorted(salaries.items(), key=lambda x: x[1], reverse=True)
        if sorted_salaries[0][1] == 0:
            return []
        richest = [sorted_salaries[0][0]]
        for i in range(1, len(sorted_salaries)):
            if sorted_salaries[i][1] != sorted_salaries[0][1]:
                break
            richest.append(sorted_salaries[i][0])
        return richest

    def add_setter(self, name) -> SetterStatus:
        try:
            pattern = self.workbook_w['PATTERN']
            setters = [pattern.cell(row=row, column=SETTER_COLUMN).value for row in range(*SETTER_ROW_RANGE)]
            for row in range(*SETTER_ROW_RANGE):
                if name in setters:
                    self.logger.info(f'Tried to add Setter({name}) which already exists')
                    return SetterStatus.ALREADY_EXISTS
                elif not pattern.cell(row=row, column=SETTER_COLUMN).value:
                    pattern.cell(row=row, column=SETTER_COLUMN).value = name
                    self.current_month_sh_w.cell(row=row, column=SETTER_COLUMN).value = name
                    break
            else:
                raise Exception
            self.workbook_w.save(self.tablepath)
            self.logger.info(f'setter "{name}" had been added to the table')
            self.__open()
            return SetterStatus.ADDED
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')
            return SetterStatus.FAILED_TO_GET

    def remove_setter(self, name) -> SetterStatus:
        try:
            pattern = self.workbook_w['PATTERN']
            for row in range(*SETTER_ROW_RANGE):
                if pattern.cell(row=row, column=SETTER_COLUMN).value == name:
                    pattern.cell(row=row, column=SETTER_COLUMN).value = None
                    self.logger.info(f'setter "{name}" had been removed from the table')
                    break
            else:
                self.logger.info(f'Routesetter with name {name} had not been found')
                return SetterStatus.NOT_FOUND
            self.workbook_w.save(self.tablepath)
            self.__open()
            return SetterStatus.REMOVED
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')
            return SetterStatus.FAILED_TO_GET

    def add_contest(self, date: str, name: str, money: int) -> ContestStatus:
        try:
            if self.contests == 0:
                column = FIRST_CONTEST_COL
            elif self.contests == 1:
                column = SECOND_CONTEST_COL
            else:
                self.logger.info('Trying to add contest info but no place for it')
                return ContestStatus.NO_PLACE
            self.current_month_sh_w.cell(row=CONTEST_PRICE_ROW, column=column).value = money
            self.current_month_sh_w.cell(row=DATE_ROW, column=column).value = date
            self.current_month_sh_w.cell(row=CONTEST_NAME_ROW, column=column).value = name
            self.workbook_w.save(self.tablepath)
            self.__open()
            self.logger.info('contest info had been added')
            return ContestStatus.ADDED
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')
            return ContestStatus.FAILED_TO_GET

    def add_contest_setter(self, setter: str) -> ContestStatus:
        try:
            if self.contests == 0:
                return ContestStatus.NO_CONTEST
            for row in range(*SETTER_ROW_RANGE):
                if self.current_month_sh_w.cell(row=row, column=SETTER_COLUMN).value == setter:
                    contest_row = row + 5
                    self.current_month_sh_w.cell(row=contest_row, column=self.contest_col).value = self.current_month_sh_w.cell(row=5, column=self.contest_col).value
                    break
            else:
                self.logger.info('Trying to add setter which is not in table')
                return ContestStatus.SETTER_ABSENT
            contest_name = self.current_month_sh_w.cell(row=3, column=self.contest_col).value
            if not self.current_month_sh_w.cell(row=contest_row, column=RESULTS_COLUMN).value:
                self.current_month_sh_w.cell(row=contest_row, column=RESULTS_COLUMN).value = contest_name
            else:
                self.current_month_sh_w.cell(row=contest_row, column=RESULTS_COLUMN).value += f' & {contest_name}'
            self.workbook_w.save(self.tablepath)
            self.__open()
            self.logger.info(f'contest participant-routesetter {setter} had been added')
            return ContestStatus.IS_CONTEST_SETTER
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')
            return ContestStatus.FAILED_TO_GET

    def remove_contest(self) -> ContestStatus:
        try:
            if self.contests == 0:
                self.logger.info('There are no contests in the table')
                return ContestStatus.NO_CONTEST
            for row in range(11, 60, 6):
                names = self.current_month_sh_w.cell(row=row, column=RESULTS_COLUMN).value
                if names:
                    if names.find(' & ') != -1:
                        names = names[:names.find(' & ')]
                    else:
                        names = None
                    self.current_month_sh_w.cell(row=row, column=RESULTS_COLUMN).value = names
                    self.current_month_sh_w.cell(row=row, column=self.contest_col).value = None

            self.current_month_sh_w.cell(row=CONTEST_NAME_ROW, column=self.contest_col).value = None
            self.current_month_sh_w.cell(row=DATE_ROW, column=self.contest_col).value = None
            self.current_month_sh_w.cell(row=CONTEST_PRICE_ROW, column=self.contest_col).value = None
            self.workbook_w.save(self.tablepath)
            self.__open()
            self.logger.info('Contest had been successfully removed')
            return ContestStatus.REMOVED
        except Exception as e:
            self.logger.error(f'Error had happened while reading Excel.\n{e}')
            return ContestStatus.FAILED_TO_GET

    def add_result(self, date: str, username: str, grade: int, amount: int) -> ResultStatus:
        for i in range(3, 13):
            table_date = self.get_cell_val(self.current_month_name, DATE_ROW, i)
            if table_date == date:
                col_ind = i
                break
        else:
            self.logger.info(f'date {date} is not in table')
            return ResultStatus.FAIL_TO_GET_DATE
        for j in range(*SETTER_ROW_RANGE):
            if self.current_month_sh_w.cell(row=j, column=SETTER_COLUMN).value == username:
                row_ind = j + grade
                self.current_month_sh_w.cell(row=row_ind, column=col_ind).value = amount
                break
        else:
            self.logger.info(f"user '{username}' voted or wrote the results but he isn't in table. His result hadn't taken")
            return ResultStatus.NOT_SETTER
        difficulty = self.current_month_sh_w.cell(row=row_ind, column=GRADES_COLUMN).value
        self.logger.info(f"{username}'s result: {difficulty}: {amount}  at {date} added successfully")
        self.workbook_w.save(self.tablepath)
        self.__open()
        return ResultStatus.ADDED
